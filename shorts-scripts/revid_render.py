#!/usr/bin/env python3
"""Submit shorts scripts from shorts-400-rewritten-v2.xlsx to the Revid AI API.

Usage:
  REVID_API_KEY=... python3 revid_render.py submit 1          # submit row 1 (post 1)
  REVID_API_KEY=... python3 revid_render.py status <pid>      # poll a render
  REVID_API_KEY=... python3 revid_render.py download <pid> out.mp4

Never hardcode the API key in this file. Pass it via the REVID_API_KEY
environment variable.

Note: endpoint paths follow Revid's public API (documented at revid.ai/api).
If Revid returns a schema error, print the full response and adjust the
payload fields; the API reports unknown/missing fields in its error body.
"""
import json
import os
import sys
import urllib.request

BASE = "https://www.revid.ai/api/public/v2"
KEY = os.environ.get("REVID_API_KEY")
XLSX = os.path.join(os.path.dirname(__file__), "shorts-400-rewritten-v2.xlsx")


def call(path, payload=None):
    req = urllib.request.Request(
        BASE + path,
        data=json.dumps(payload).encode() if payload is not None else None,
        headers={"key": KEY, "Content-Type": "application/json"},
        method="POST" if payload is not None else "GET",
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read().decode())


def load_row(row_number):
    import openpyxl  # pip install openpyxl
    ws = openpyxl.load_workbook(XLSX).active
    hdr = [c.value for c in ws[1]]
    row = [c.value for c in ws[row_number + 1]]
    return dict(zip(hdr, row))


def submit(row_number):
    r = load_row(row_number)
    payload = {
        "webhook": None,
        "creationParams": {
            "inputText": r["script_45_to_55_seconds"],
            "flow": "text-to-video",
            "targetDuration": 55,
            "ratio": "9 / 16",
            "mediaType": "stockVideo",
            "hasToGenerateVoice": True,
            "hasToSearchMedia": True,
            "captionPresetName": "Wrap 1",
            "hasToGenerateCaptions": True,
            "searchTerms": ["hotel lobby", "hotel reception", "Dubai"],
        },
    }
    try:
        resp = call("/render", payload)
    except urllib.error.HTTPError as e:
        print("HTTP", e.code, e.read().decode())
        sys.exit(1)
    print(json.dumps(resp, indent=2))


def status(pid):
    try:
        resp = call(f"/status?pid={pid}")
    except urllib.error.HTTPError as e:
        print("HTTP", e.code, e.read().decode())
        sys.exit(1)
    print(json.dumps(resp, indent=2))


if __name__ == "__main__":
    if not KEY:
        sys.exit("Set REVID_API_KEY first.")
    cmd = sys.argv[1] if len(sys.argv) > 1 else "submit"
    if cmd == "submit":
        submit(int(sys.argv[2]) if len(sys.argv) > 2 else 1)
    elif cmd == "status":
        status(sys.argv[2])
    else:
        sys.exit(f"Unknown command {cmd}")
